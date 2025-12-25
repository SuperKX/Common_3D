import pandas as pd
import numpy as np
import os
# 导入data_dict和label_dict
from application.PTV3_dataprocess import data_dict, label_dict
import points.points_io as points_io
import points.points_eval as points_eval

def generate_scene_group_csv(output_csv=None, data_version=None, data_folder=None, label_version='label05_V1',
                             pred_folder=None, label_version_pred='label05_V1'):
    '''
    根据data_dict.py中的data_version20240905创建场景分组CSV表格
    todo： 输入改成两个标签值的方式。外部套壳，支持单文件、多文件输入。
    参数:
        output_csv: 输出CSV文件路径（可选）
        data_version: 数据版本字典（可选，默认使用data_version20240905）
        data_folder: 数据文件夹路径（可选，用于统计点云数量）
        label_version: 标签版本（可选，默认使用label05_V1）
        pred_folder: 推理输出数据文件夹路径（可选，用于计算评估指标）
        label_version_pred: 推理数据标签版本（可选，默认使用label05_V1）

    返回:
        DataFrame包含分组和场景名信息，以及点云数量统计（如果提供data_folder）
        和评估指标（如果提供pred_folder）
    '''
    # 如果没有指定data_version，使用默认版本
    if data_version is None:
        data_version = data_dict.data_version20240905

    # 获取标签定义
    if label_version in label_dict.LabelRegistry.label_def:
        label_def = label_dict.LabelRegistry.label_def[label_version]
        # 根据标签定义生成类别列名
        category_columns = [label_def[i] for i in sorted(label_def.keys())]
        class_num = len(label_def)
    else:
        raise ValueError(f"未找到标签版本: {label_version}")

    print(f"\n生成场景分组CSV表格...")
    print(f"数据版本: data_version20240905")
    print(f"标签版本: {label_version}")
    if data_folder:
        print(f"数据文件夹: {data_folder}")
    if pred_folder:
        print(f"推理文件夹: {pred_folder}")
        print(f"推理标签版本: {label_version_pred}")

    # 存储所有数据
    all_data = []

    # 各类别数量统计
    class_sum = {'train':{}, 'val':{}, 'test':{}}
    for cat_name in category_columns:
        class_sum['train'][cat_name] = 0
        class_sum['val'][cat_name] = 0
        class_sum['test'][cat_name] = 0
    class_sum['train']['分组点云总数'] = 0
    class_sum['val']['分组点云总数'] = 0
    class_sum['test']['分组点云总数'] = 0


    # points_num_class = {}
    # label_def

    # 1 遍历每个分组
    for group_name, scene_set in data_version.items():
        # 遍历该分组中的每个场景
        for scene_name in scene_set:
            # 1.1 场景基本信息
            row_data = {
                '分组': group_name,
                '场景名': scene_name
            }
            if data_folder:  # 如果提供了数据文件夹，统计点云数量
                ply_path = os.path.join(data_folder, f"{scene_name}.ply")

                if not os.path.exists(ply_path):
                    raise FileNotFoundError(f"未找到点云文件 {ply_path}")
                points_dict = points_io.parse_cloud_to_dict(ply_path)
                label_version_plystyle = 'label_' + label_version  # ply 文件前缀
                if not label_version_plystyle in points_dict:
                    raise ValueError(f"{scene_name}.ply 中没有{label_version}标签信息")
                labels = points_dict[label_version_plystyle]
                # points_num_info = points_eval.file_label_info(ply_path, label_version)

                # 1.2 点云数量
                row_data['点云数量'] = len(labels)
                label_counts = {}
                for label_id, label_name in label_def.items():
                    count = np.sum(labels == label_id)
                    label_counts[label_name] = count
                for cat_name in category_columns:  # 类别数量
                    row_data[cat_name] = label_counts.get(cat_name, 0)
                    class_sum[group_name][cat_name] += row_data[cat_name]   # 各类别数量占比
                # 1.3 点云内数据占比
                row_data['数据占比'] = 1
                for cat_name in category_columns:  # 类别比例
                    row_data['drt_'+cat_name] = row_data[cat_name]/row_data['点云数量']

            all_data.append(row_data)
    # 1.4 计算分组内类别比例
    for cat_name in category_columns:
        class_sum['train']['分组点云总数'] += class_sum['train'][cat_name]
        class_sum['val']['分组点云总数'] += class_sum['val'][cat_name]
        class_sum['test']['分组点云总数'] += class_sum['test'][cat_name]
    for data_row in all_data:
        data_row['分组占比-总点云'] = data_row['点云数量']/class_sum[data_row['分组']]['分组点云总数']
        for cat_name in category_columns:
            data_row['grt_' + cat_name] = data_row[cat_name] / class_sum[data_row['分组']][cat_name]

    # 1.5 如果提供了推理文件夹，计算评估指标
    if pred_folder:
        for data_row in all_data:
            scene_name = data_row['场景名']
            label_path = os.path.join(data_folder, f"{scene_name}.ply")
            pred_path = os.path.join(pred_folder, f"{scene_name}.ply")

            if not os.path.exists(pred_path):
                print(f"警告: 未找到推理文件 {pred_path}，跳过评估")
                data_row['平均召回率'] = np.nan
                data_row['平均精确率'] = np.nan
                data_row['平均iou'] = np.nan
                # 各类别指标设为NaN
                for cat_name in category_columns:
                    data_row[f'{cat_name}_recall'] = np.nan
                    data_row[f'{cat_name}_precision'] = np.nan
                    data_row[f'{cat_name}_iou'] = np.nan
                # grass混淆度设为NaN
                data_row['grass2background'] = np.nan
                data_row['grass2vegetation'] = np.nan
                data_row['background2grass'] = np.nan
                data_row['vegetation2grass'] = np.nan
                # grass优先级评分设为NaN
                data_row['优先级评分'] = np.nan
                data_row['grass_混淆程度评分'] = np.nan
            else:
                # 读取标签数据
                labeled_dict = points_io.parse_cloud_to_dict(label_path)
                label_version_plystyle = 'label_' + label_version
                if not label_version_plystyle in labeled_dict:
                    raise ValueError(f"{scene_name}.ply 中没有{label_version}标签信息")
                labels = labeled_dict[label_version_plystyle]
                # 读取推理数据
                pred_dict = points_io.parse_cloud_to_dict(pred_path)
                label_version_pred_plystyle = 'label_' + label_version_pred
                if not label_version_pred_plystyle in pred_dict:
                    raise ValueError(f"{scene_name}.ply 中没有{label_version_pred}标签信息")
                labels_pred = pred_dict[label_version_pred_plystyle]

                # 计算混淆矩阵
                score_matrix = points_eval.matrix_eval(labels, labels_pred, class_num)

                # 计算评估指标
                score_rec, score_pre, score_iou = points_eval.eval_result(score_matrix, class_num)

                # 计算平均指标
                data_row['平均召回率'] = np.mean(score_rec)
                data_row['平均精确率'] = np.mean(score_pre)
                data_row['平均iou'] = np.mean(score_iou)

                # 各类别指标
                for i, cat_name in enumerate(category_columns):
                    data_row[f'{cat_name}_recall'] = score_rec[i]
                    data_row[f'{cat_name}_precision'] = score_pre[i]
                    data_row[f'{cat_name}_iou'] = score_iou[i]

                # 1.6 混淆度计算（grass与background、vegetation的混淆比例）
                # 获取grass、background、vegetation的标签ID
                grass_id = background_id = vegetation_id = None
                for label_id, label_name in label_def.items():
                    if label_name == 'grass':
                        grass_id = label_id
                    elif label_name == 'background':
                        background_id = label_id
                    elif label_name == 'vegetation':
                        vegetation_id = label_id

                if grass_id is not None and background_id is not None and vegetation_id is not None:
                    # grass被误分类为background的比例
                    grass_true_count = np.sum(score_matrix[grass_id, :])
                    if grass_true_count > 0:
                        data_row['grass2background'] = score_matrix[grass_id][background_id] / grass_true_count
                    else:
                        data_row['grass2background'] = np.nan

                    # grass被误分类为vegetation的比例
                    if grass_true_count > 0:
                        data_row['grass2vegetation'] = score_matrix[grass_id][vegetation_id] / grass_true_count
                    else:
                        data_row['grass2vegetation'] = np.nan

                    # background被误分类为grass的比例
                    background_true_count = np.sum(score_matrix[background_id, :])
                    if background_true_count > 0:
                        data_row['background2grass'] = score_matrix[background_id][grass_id] / background_true_count
                    else:
                        data_row['background2grass'] = np.nan

                    # vegetation被误分类为grass的比例
                    vegetation_true_count = np.sum(score_matrix[vegetation_id, :])
                    if vegetation_true_count > 0:
                        data_row['vegetation2grass'] = score_matrix[vegetation_id][grass_id] / vegetation_true_count
                    else:
                        data_row['vegetation2grass'] = np.nan

                # 1.7 grass优先级评分计算
                # 获取grass的分组占比（grt_grass）和grass_iou
                grt_grass = data_row.get('grt_grass', np.nan)
                grass_iou = data_row.get('grass_iou', np.nan)

                # 优先级评分 = grt_grass * (1 - grass_iou)
                if not np.isnan(grt_grass) and not np.isnan(grass_iou):
                    data_row['优先级评分'] = grt_grass * (1 - grass_iou)
                else:
                    data_row['优先级评分'] = np.nan

                # grass_混淆程度评分 = grt_grass * max(混淆度)
                if not np.isnan(grt_grass):
                    confusion_values = [
                        data_row.get('grass2background', np.nan),
                        data_row.get('grass2vegetation', np.nan),
                        data_row.get('background2grass', np.nan),
                        data_row.get('vegetation2grass', np.nan)
                    ]
                    # 过滤掉NaN值
                    valid_confusions = [v for v in confusion_values if not np.isnan(v)]
                    if valid_confusions:
                        data_row['grass_混淆程度评分'] = grt_grass * max(valid_confusions)
                    else:
                        data_row['grass_混淆程度评分'] = np.nan
                else:
                    data_row['grass_混淆程度评分'] = np.nan

    # 转换为DataFrame
    df = pd.DataFrame(all_data)

    # 2 按分组排序，然后按场景名排序
    if len(df) > 0:
        # 按分组顺序排序（train -> val -> test）
        group_order = {'train': 0, 'val': 1, 'test': 2}
        df['_group_order'] = df['分组'].map(group_order)
        df = df.sort_values(['_group_order', '场景名'])  # 优先级排序，分组>场景名
        df = df.drop('_group_order', axis=1)
        df = df.reset_index(drop=True)

    # 3 保存到CSV
    if output_csv and len(df) > 0:
        df.to_csv(output_csv, index=False, encoding='utf-8-sig')
        print(f"\n场景分组CSV已保存到: {output_csv}")

    # 打印统计信息
    if len(df) > 0:
        print("\n分组统计:")
        for group_name in data_version.keys():
            count = len(data_version[group_name])
            print(f"  {group_name}: {count} 个场景")
        print(f"  总计: {len(df)} 个场景")
    return df


def csv_to_xlsx(csv_path, xlsx_path=None, columns_dict=None):
    '''
    读入CSV表格，并将标题加粗，特定列转换为百分比格式，写出xlsx格式

    参数:
        csv_path: 输入CSV文件路径
        xlsx_path: 输出xlsx文件路径（可选，默认与csv_path同目录，扩展名改为xlsx）
        columns_dict: 属性字典，格式为 {'属性组名': ['列名1', '列名2', ...]}
                      如果为None，则输出所有属性；否则只输出字典中指定的属性

    返回:
        None
    '''
    # 如果没有指定xlsx_path，默认使用同目录下同文件名的xlsx
    if xlsx_path is None:
        xlsx_path = os.path.splitext(csv_path)[0] + '.xlsx'

    # 读取CSV文件
    df = pd.read_csv(csv_path, encoding='utf-8-sig')

    # 定义完整的属性组
    full_attribute_groups = {
        '点云统计': ['点云数量', 'background', 'building', 'car', 'vegetation', 'grass'],
        '场景比例': ['数据占比', 'drt_background', 'drt_building', 'drt_car', 'drt_vegetation', 'drt_grass'],
        '分组类别比例': ['分组占比-总点云', 'grt_background', 'grt_building', 'grt_car', 'grt_vegetation', 'grt_grass'],
        '精度统计': ['平均召回率', '平均精确率', '平均iou',
                     'background_recall', 'background_precision', 'background_iou',
                     'building_recall', 'building_precision', 'building_iou',
                     'car_recall', 'car_precision', 'car_iou',
                     'vegetation_recall', 'vegetation_precision', 'vegetation_iou',
                     'grass_recall', 'grass_precision', 'grass_iou'],
        '混淆对比': ['grass2background', 'grass2vegetation', 'background2grass', 'vegetation2grass'],
        '评分': ['优先级评分', 'grass_混淆程度评分']
    }

    # 根据columns_dict筛选需要输出的列
    if columns_dict is not None:
        # 收集所有需要输出的列
        output_columns = []
        for group_name in columns_dict.keys():
            if group_name in full_attribute_groups:
                # 对于指定的属性组，只输出columns_dict中指定的列（存在于数据中的）
                for col in columns_dict[group_name]:
                    if col in df.columns:
                        output_columns.append(col)
        # 添加'分组'和'场景名'列（如果存在）
        if '分组' in df.columns and '分组' not in output_columns:
            output_columns.insert(0, '分组')
        if '场景名' in df.columns and '场景名' not in output_columns:
            if '分组' in output_columns:
                output_columns.insert(1, '场景名')
            else:
                output_columns.insert(0, '场景名')
        # 筛选DataFrame
        df = df[output_columns]
        # 根据筛选后的列更新属性组
        attribute_groups = {}
        for group_name in columns_dict.keys():
            if group_name in full_attribute_groups:
                # 只保留存在于输出数据中的列
                valid_columns = [col for col in columns_dict[group_name] if col in df.columns]
                if valid_columns:
                    attribute_groups[group_name] = valid_columns
    else:
        # 使用完整的属性组
        attribute_groups = full_attribute_groups

    # 需要转换为百分比的列名
    percent_columns = [
        'drt_background', 'drt_building', 'drt_car', 'drt_vegetation', 'drt_grass',
        '分组占比-总点云',
        'grt_background', 'grt_building', 'grt_car', 'grt_vegetation', 'grt_grass',
        '平均召回率', '平均精确率', '平均iou',
        'background_recall', 'background_precision', 'background_iou',
        'building_recall', 'building_precision', 'building_iou',
        'car_recall', 'car_precision', 'car_iou',
        'vegetation_recall', 'vegetation_precision', 'vegetation_iou',
        'grass_recall', 'grass_precision', 'grass_iou',
        'grass2background', 'grass2vegetation', 'background2grass', 'vegetation2grass',
        '优先级评分', 'grass_混淆程度评分'
    ]

    # 写入xlsx文件并设置格式
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']

        # 导入样式模块
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

        # 定义标题填充色
        color_map = {
            # 浅绿色：点云数量和各类别数量
            '点云数量': 'C6E0B4',
            'background': 'C6E0B4',
            'building': 'C6E0B4',
            'car': 'C6E0B4',
            'vegetation': 'C6E0B4',
            'grass': 'C6E0B4',
            # 浅黄色：数据占比和各类别数据占比
            '数据占比': 'FFF2CC',
            'drt_background': 'FFF2CC',
            'drt_building': 'FFF2CC',
            'drt_car': 'FFF2CC',
            'drt_vegetation': 'FFF2CC',
            'drt_grass': 'FFF2CC',
            # 浅橙色：分组占比和各类别分组占比
            '分组占比-总点云': 'FFE0B2',
            'grt_background': 'FFE0B2',
            'grt_building': 'FFE0B2',
            'grt_car': 'FFE0B2',
            'grt_vegetation': 'FFE0B2',
            'grt_grass': 'FFE0B2',
            # 浅红色：评估指标
            '平均召回率': 'F4CCCC',
            '平均精确率': 'F4CCCC',
            '平均iou': 'F4CCCC',
            'background_recall': 'F4CCCC',
            'background_precision': 'F4CCCC',
            'background_iou': 'F4CCCC',
            'building_recall': 'F4CCCC',
            'building_precision': 'F4CCCC',
            'building_iou': 'F4CCCC',
            'car_recall': 'F4CCCC',
            'car_precision': 'F4CCCC',
            'car_iou': 'F4CCCC',
            'vegetation_recall': 'F4CCCC',
            'vegetation_precision': 'F4CCCC',
            'vegetation_iou': 'F4CCCC',
            'grass_recall': 'F4CCCC',
            'grass_precision': 'F4CCCC',
            'grass_iou': 'F4CCCC',
            # 浅紫色：grass混淆度
            'grass2background': 'D9D2E9',
            'grass2vegetation': 'D9D2E9',
            'background2grass': 'D9D2E9',
            'vegetation2grass': 'D9D2E9',
            # 浅蓝色：评分指标
            '优先级评分': 'CFE2F3',
            'grass_混淆程度评分': 'CFE2F3',
        }

        # 定义边框样式（黑色细边框）
        left_border = Border(left=Side(style='thin', color='000000'))
        right_border = Border(right=Side(style='thin', color='000000'))

        # 获取每个属性组的第一列和最后一列
        group_first_columns = set()  # 需要左边框的列
        group_last_columns = set()   # 需要右边框的列
        for group_name, columns in attribute_groups.items():
            group_first_columns.add(columns[0])  # 第一列
            group_last_columns.add(columns[-1])   # 最后一列

        # 设置标题行加粗和填充色，以及属性组边框
        for cell in worksheet[1]:  # 第一行是标题
            cell.font = Font(bold=True)
            if cell.value in color_map:
                cell.fill = PatternFill(start_color=color_map[cell.value], end_color=color_map[cell.value], fill_type='solid')
            # 设置属性组左边框
            if cell.value in group_first_columns:
                cell.border = left_border
            # 设置属性组右边框
            elif cell.value in group_last_columns:
                cell.border = right_border

        # 为所有数据行设置属性组边框和"分组"/"场景名"列格式
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                column_name = worksheet.cell(1, cell.column).value

                # 设置"分组"和"场景名"列加粗
                if column_name in ['分组', '场景名']:
                    cell.font = Font(bold=True)

                # 设置"分组"列的填充色
                if column_name == '分组' and cell.value in ['train', 'val', 'test']:
                    if cell.value == 'train':
                        cell.fill = PatternFill(start_color='CFE2F3', end_color='CFE2F3', fill_type='solid')
                    elif cell.value == 'val':
                        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    elif cell.value == 'test':
                        cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')

                # 设置属性组边框
                if column_name in group_first_columns:
                    cell.border = left_border
                elif column_name in group_last_columns:
                    cell.border = right_border

        # 设置百分比列的格式
        for row in worksheet.iter_rows(min_row=2):  # 从第二行开始（数据行）
            for cell in row:
                column_name = worksheet.cell(1, cell.column).value  # 获取列名
                if column_name in percent_columns:
                    if isinstance(cell.value, (int, float)) and not np.isnan(cell.value):
                        # 转换为百分比格式（保留两位小数）
                        cell.number_format = FORMAT_PERCENTAGE_00

                        # 分组占比列的条件格式化
                        if column_name in ['分组占比-总点云', 'grt_background', 'grt_building', 'grt_car', 'grt_vegetation', 'grt_grass']:
                            value = cell.value
                            if value > 0.2:
                                # 超过20%：红色加粗
                                cell.font = Font(bold=True, color='FF0000')
                            elif value >= 0.1:
                                # 10%-20%：橙色加粗
                                cell.font = Font(bold=True, color='FFA500')
                            elif value >= 0.05:
                                # 5%-10%：紫色加粗
                                cell.font = Font(bold=True, color='800080')

                        # grass混淆度列的条件格式化
                        elif column_name in ['grass2background', 'grass2vegetation', 'background2grass', 'vegetation2grass']:
                            value = cell.value
                            if value > 0.5:
                                # 超过50%：红色加粗
                                cell.font = Font(bold=True, color='FF0000')
                            elif value >= 0.2:
                                # 20%-50%：橙色加粗
                                cell.font = Font(bold=True, color='FFA500')
                            elif value >= 0.1:
                                # 10%-20%：紫色加粗
                                cell.font = Font(bold=True, color='800080')

                        # grass_iou的条件格式化
                        elif column_name == 'grass_iou':
                            value = cell.value
                            if value < 0.2:
                                # 小于20%：红色加粗
                                cell.font = Font(bold=True, color='FF0000')
                            elif value < 0.4:
                                # 20%-40%：橙色加粗
                                cell.font = Font(bold=True, color='FFA500')
                            elif value < 0.6:
                                # 40%-60%：紫色加粗
                                cell.font = Font(bold=True, color='800080')

                        # 优先级评分和grass混淆程度评分的条件格式化
                        elif column_name in ['优先级评分', 'grass_混淆程度评分']:
                            value = cell.value
                            if value > 0.1:
                                # 超过10%：红色加粗
                                cell.font = Font(bold=True, color='FF0000')
                            elif value >= 0.05:
                                # 5%-10%：橙色加粗
                                cell.font = Font(bold=True, color='FFA500')
                            elif value >= 0.02:
                                # 2%-5%：紫色加粗
                                cell.font = Font(bold=True, color='800080')

        # 冻结前两列和第一行
        worksheet.freeze_panes = "C2"

    print(f"\nXLSX文件已保存到: {xlsx_path}")


if __name__ == '__main__':
    if False:  # 数据分析
        output_csv = r'H:\commonFunc_3D\application\PTV3_dataprocess/输出测试.csv'
        label_folder = r'J:\DATASET\BIMTwins\版本备份\多标签_动态维护版'
        pred_folder = r"H:\TempProcess\20251220数据传输\2合并标签"
        label_version_pred = "label05_V1_pred"
        # 注意，内部会默认转ply风格的字典，不需要外部添加（待确定）
        generate_scene_group_csv(output_csv, data_folder=label_folder, pred_folder=pred_folder,
                                 label_version_pred=label_version_pred)

    if True:  # 数据格式优化
        input_csv = r'H:\commonFunc_3D\application\PTV3_dataprocess/输出测试.csv'
        output_xlsx = r'H:\commonFunc_3D\application\PTV3_dataprocess/输出测试.xlsx'
#         columns_dict = {
#             "点云统计": [
#                 "点云数量", "background", "building", "car", "vegetation", "grass"
#             ],
#             "场景比例": [
#                 "数据占比", "drt_background", "drt_building", "drt_car", "drt_vegetation", "drt_grass"
#             ],
#             "分组类别比例": [
#                 "分组占比-总点云", "grt_background", "grt_building", "grt_car", "grt_vegetation", "grt_grass"
#             ],
#             "精度统计": [
#                 "平均召回率", "平均精确率", "平均iou", "background_recall", "background_precision",
#                 "background_iou", "building_recall", "building_precision", "building_iou",
#                 "car_recall", "car_precision", "car_iou", "vegetation_recall", "vegetation_precision",
#                 "vegetation_iou", "grass_recall", "grass_precision", "grass_iou"
#             ],
#             "混淆对比": [
#                 "grass2background", "grass2vegetation", "background2grass", "vegetation2grass"
#             ],
        #             "评分": ["优先级评分", "grass_混淆程度评分"]
#         }
        columns_dict = {
            "点云统计": [
                "点云数量", "background", "building", "car", "vegetation", "grass"
            ],
            "场景比例": [
                 "drt_background", "drt_building", "drt_car", "drt_vegetation", "drt_grass"
            ],
            "分组类别比例": [
                "分组占比-总点云", "grt_background", "grt_building", "grt_car", "grt_vegetation", "grt_grass"
            ],
            "精度统计": [
                "平均召回率", "平均精确率", "平均iou", "grass_recall", "grass_precision", "grass_iou"
            ],
            "混淆对比": [
                "grass2background", "grass2vegetation", "background2grass", "vegetation2grass"
            ],
            "评分": ["优先级评分", "grass_混淆程度评分"]
        }

        csv_to_xlsx(input_csv, output_xlsx, columns_dict)

